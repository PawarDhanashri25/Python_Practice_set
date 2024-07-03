import openpyxl

wb= openpyxl.load_workbook('student.xlsx')  # this will store worksheet object

print(wb.sheetnames)

for sheet in wb:
    print(sheet)   # sheet name
    print(sheet.title)  # sheet title


sheet = wb['Student_data']

print( '------------------Reading Cells from sheet---------------------')
b2_cell= sheet['B2']
print(b2_cell.value)

print(sheet.cell(row=3, column=2).value)        # reading tha particular cell data

print(b2_cell.row, b2_cell.column)     # reading the row and column number

print("Print the data type")

print(sheet['A2'].data_type)

print(sheet['A4'].encoding)
print(sheet['b4'].encoding)     # print the encoding scheme

print(sheet['D4'].parent)

print(dir(b2_cell))    # print all the attributes and methods

cell_range = sheet['B2' : 'C5']

for name, roll in cell_range:
    print(f'Roll No: {roll.value} Name: {name.value}')


print(f'Sheet Dimensions: {sheet.dimensions}')

print(sheet.max_row, sheet.max_column)

for a, b, c, d in sheet[sheet.dimensions]:
    print(a.value, b.value, c.value, d.value)

# printing the excel data
print('------------------printing excel data -----------------------')

for row in sheet.rows:
    for cell in row:
        print(f'{cell.value}', end='')
    print('\n')

for row in sheet.values:
    print(row)
    






